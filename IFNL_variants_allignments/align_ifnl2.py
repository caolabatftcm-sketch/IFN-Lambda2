"""
Needleman-Wunsch global alignment of human vs mouse IFN-Lambda 2
with BLOSUM62 scoring, gap open=-10, gap extend=-0.5
Outputs: Excel residue mapping + HTML alignment view
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os

# ── Sequences ────────────────────────────────────────────────────────────────
HUMAN = "MKLDMTGDCTPVLVLMAAVLTVTGAVPVARLHGALPDARGCHIAQFKSLSPQELQAFKRAKDALEESLLLKDCRCHSRLFPRTWDLRQLQVRERPMALEAELALTLKVLEATADTDPALVDVLDQPLHTLHHILSQFRACIQPQPTAGPRTRGRLHHWLYRLQEAPKKESPGCLEASVTFNLFRLLTRDLNCVASGDLCV"
MOUSE = "MLLLLLPLLLAAVLTRTQADPVPRATRLPVEAKDCHIAQFKSLSPKELQAFKKAKDAIEKRLLEKDLRCSSHLFPRAWDLKQLQVQERPKALQAEVALTLKVWENMTDSALATILGQPLHTLSHIHSQLQTCTQLQATAEPRSPSRRLSRWLHRLQEAQSKETPGCLEASVTSNLFRLLTRDLKCVANGDQCV"

# ── BLOSUM62 ─────────────────────────────────────────────────────────────────
BLOSUM62_STR = """
   A  R  N  D  C  Q  E  G  H  I  L  K  M  F  P  S  T  W  Y  V  B  Z  X  *
A  4 -1 -2 -2  0 -1 -1  0 -2 -1 -1 -1 -1 -2 -1  1  0 -3 -2  0 -2 -1  0 -4
R -1  5  0 -2 -3  1  0 -2  0 -3 -2  2 -1 -3 -2 -1 -1 -3 -2 -3 -1  0 -1 -4
N -2  0  6  1 -3  0  0  0  1 -3 -3  0 -2 -3 -2  1  0 -4 -2 -3  3  0 -1 -4
D -2 -2  1  6 -3  0  2 -1 -1 -3 -4 -1 -3 -3 -1  0 -1 -4 -3 -3  4  1 -1 -4
C  0 -3 -3 -3  9 -3 -4 -3 -3 -1 -1 -3 -1 -2 -3 -1 -1 -2 -2 -1 -3 -3 -2 -4
Q -1  1  0  0 -3  5  2 -2  0 -3 -2  1  0 -3 -1  0 -1 -2 -1 -2  0  3 -1 -4
E -1  0  0  2 -4  2  5 -2  0 -3 -3  1 -2 -3 -1  0 -1 -3 -2 -2  1  4 -1 -4
G  0 -2  0 -1 -3 -2 -2  6 -2 -4 -4 -2 -3 -3 -2  0 -2 -2 -3 -3 -1 -2 -1 -4
H -2  0  1 -1 -3  0  0 -2  8 -3 -3 -1 -2 -1 -2 -1 -2 -2  2 -3  0  0 -1 -4
I -1 -3 -3 -3 -1 -3 -3 -4 -3  4  2 -3  1  0 -3 -2 -1 -3 -1  3 -3 -3 -1 -4
L -1 -2 -3 -4 -1 -2 -3 -4 -3  2  4 -2  2  0 -3 -2 -1 -2 -1  1 -4 -3 -1 -4
K -1  2  0 -1 -3  1  1 -2 -1 -3 -2  5 -1 -3 -1  0 -1 -3 -2 -2  0  1 -1 -4
M -1 -1 -2 -3 -1  0 -2 -3 -2  1  2 -1  5  0 -2 -1 -1 -1 -1  1 -3 -1 -1 -4
F -2 -3 -3 -3 -2 -3 -3 -3 -1  0  0 -3  0  6 -4 -2 -2  1  3 -1 -3 -3 -1 -4
P -1 -2 -2 -1 -3 -1 -1 -2 -2 -3 -3 -1 -2 -4  7 -1 -1 -4 -3 -2 -2 -1 -2 -4
S  1 -1  1  0 -1  0  0  0 -1 -2 -2  0 -1 -2 -1  4  1 -3 -2 -2  0  0  0 -4
T  0 -1  0 -1 -1 -1 -1 -2 -2 -1 -1 -1 -1 -2 -1  1  5 -2 -2  0 -1 -1  0 -4
W -3 -3 -4 -4 -2 -2 -3 -2 -2 -3 -2 -3 -1  1 -4 -3 -2 11  2 -3 -4 -3 -2 -4
Y -2 -2 -2 -3 -2 -1 -2 -3  2 -1 -1 -2 -1  3 -3 -2 -2  2  7 -1 -3 -2 -1 -4
V  0 -3 -3 -3 -1 -2 -2 -3 -3  3  1 -2  1 -1 -2 -2  0 -3 -1  4 -3 -2 -1 -4
B -2 -1  3  4 -3  0  1 -1  0 -3 -4  0 -3 -3 -2  0 -1 -4 -3 -3  4  1 -1 -4
Z -1  0  0  1 -3  3  4 -2  0 -3 -3  1 -1 -3 -1  0 -1 -3 -2 -2  1  4 -1 -4
X  0 -1 -1 -1 -2 -1 -1 -1 -1 -1 -1 -1 -1 -1 -2  0  0 -2 -1 -1 -1 -1 -1 -4
* -4 -4 -4 -4 -4 -4 -4 -4 -4 -4 -4 -4 -4 -4 -4 -4 -4 -4 -4 -4 -4 -4 -4  1
"""

def parse_blosum62(text):
    lines = [l for l in text.strip().split('\n') if l.strip()]
    aa_order = lines[0].split()
    matrix = {}
    for line in lines[1:]:
        parts = line.split()
        aa1 = parts[0]
        for j, aa2 in enumerate(aa_order):
            matrix[(aa1, aa2)] = int(parts[j + 1])
    return matrix

BLOSUM62 = parse_blosum62(BLOSUM62_STR)

def blosum_score(a, b):
    key = (a.upper(), b.upper())
    rkey = (b.upper(), a.upper())
    return BLOSUM62.get(key, BLOSUM62.get(rkey, -1))

# ── Needleman-Wunsch (affine gap) ────────────────────────────────────────────
def needleman_wunsch(seq1, seq2, gap_open=-10, gap_extend=-0.5):
    m, n = len(seq1), len(seq2)
    NEG_INF = float('-inf')

    # Three matrices: M (match), IX (gap in seq2/human), IY (gap in seq1/mouse)
    M  = [[NEG_INF]*(n+1) for _ in range(m+1)]
    IX = [[NEG_INF]*(n+1) for _ in range(m+1)]
    IY = [[NEG_INF]*(n+1) for _ in range(m+1)]

    M[0][0] = 0
    for i in range(1, m+1):
        IX[i][0] = gap_open + (i-1)*gap_extend
        M[i][0]  = gap_open + (i-1)*gap_extend
    for j in range(1, n+1):
        IY[0][j] = gap_open + (j-1)*gap_extend
        M[0][j]  = gap_open + (j-1)*gap_extend

    for i in range(1, m+1):
        for j in range(1, n+1):
            s = blosum_score(seq1[i-1], seq2[j-1])
            M[i][j]  = s + max(M[i-1][j-1], IX[i-1][j-1], IY[i-1][j-1])
            IX[i][j] = max(M[i-1][j] + gap_open, IX[i-1][j] + gap_extend)
            IY[i][j] = max(M[i][j-1] + gap_open, IY[i][j-1] + gap_extend)

    # Traceback
    align1, align2 = [], []
    i, j = m, n
    state = 'M'  # current matrix

    while i > 0 or j > 0:
        if state == 'M':
            if i == 0:
                align1.append('-'); align2.append(seq2[j-1]); j -= 1
            elif j == 0:
                align1.append(seq1[i-1]); align2.append('-'); i -= 1
            else:
                s = blosum_score(seq1[i-1], seq2[j-1])
                prev_M  = M[i-1][j-1]
                prev_IX = IX[i-1][j-1]
                prev_IY = IY[i-1][j-1]
                best = max(prev_M, prev_IX, prev_IY)
                align1.append(seq1[i-1]); align2.append(seq2[j-1])
                i -= 1; j -= 1
                if best == prev_IX: state = 'IX'
                elif best == prev_IY: state = 'IY'
        elif state == 'IX':
            align1.append(seq1[i-1]); align2.append('-')
            if M[i-1][j] + gap_open >= IX[i-1][j] + gap_extend:
                state = 'M'
            i -= 1
        elif state == 'IY':
            align1.append('-'); align2.append(seq2[j-1])
            if M[i][j-1] + gap_open >= IY[i][j-1] + gap_extend:
                state = 'M'
            j -= 1

    align1 = ''.join(reversed(align1))
    align2 = ''.join(reversed(align2))
    return align1, align2

# ── Run alignment ────────────────────────────────────────────────────────────
print("Running Needleman-Wunsch alignment...")
aln_human, aln_mouse = needleman_wunsch(HUMAN, MOUSE)
assert len(aln_human) == len(aln_mouse), "Alignment length mismatch!"

# Build mapping table
mapping = []
h_pos = 0  # 1-based human residue counter
m_pos = 0  # 1-based mouse residue counter

for col_i, (h, m) in enumerate(zip(aln_human, aln_mouse)):
    h_aa = h if h != '-' else None
    m_aa = m if m != '-' else None
    if h_aa: h_pos += 1
    if m_aa: m_pos += 1

    if h_aa and m_aa:
        identical = (h_aa.upper() == m_aa.upper())
        s = blosum_score(h_aa, m_aa)
        conservative = (not identical) and (s > 0)
    else:
        identical = False
        conservative = False
        s = None

    mapping.append({
        'aln_col':     col_i + 1,
        'h_pos':       h_pos if h_aa else None,
        'h_aa':        h_aa,
        'm_pos':       m_pos if m_aa else None,
        'm_aa':        m_aa,
        'identity':    identical,
        'conservative':conservative,
        'blosum62':    s,
    })

# Stats
paired = [r for r in mapping if r['h_aa'] and r['m_aa']]
ident  = sum(1 for r in paired if r['identity'])
cons   = sum(1 for r in paired if r['conservative'])
gaps   = sum(1 for r in mapping if not r['h_aa'] or not r['m_aa'])

pct_id   = 100 * ident / len(paired) if paired else 0
pct_sim  = 100 * (ident + cons) / len(paired) if paired else 0
aln_len  = len(aln_human)

print(f"Alignment length : {aln_len}")
print(f"Identical        : {ident}/{len(paired)} ({pct_id:.1f}%)")
print(f"Similar (+ BLSM) : {ident+cons}/{len(paired)} ({pct_sim:.1f}%)")
print(f"Gap columns      : {gaps}")

# ── Excel output ─────────────────────────────────────────────────────────────
OUT_XLSX = "/sessions/affectionate-zealous-hypatia/mnt/outputs/IFNL2_human_mouse_mapping.xlsx"
os.makedirs(os.path.dirname(OUT_XLSX), exist_ok=True)

wb = openpyxl.Workbook()

# ─ Sheet 1: Residue Mapping ─
ws = wb.active
ws.title = "Residue Mapping"

# Color palette
HDR_FILL   = PatternFill("solid", start_color="2E4057")
IDENT_FILL = PatternFill("solid", start_color="D4EDDA")
CONS_FILL  = PatternFill("solid", start_color="FFF3CD")
GAP_FILL   = PatternFill("solid", start_color="F8D7DA")
NONE_FILL  = PatternFill("solid", start_color="FFFFFF")
ALT_FILL   = PatternFill("solid", start_color="F5F5F5")

HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Arial", size=10)
CENTER = Alignment(horizontal="center", vertical="center")

headers = [
    "Aln Col", "Human Pos", "Human AA", "Mouse Pos", "Mouse AA",
    "Conservation", "BLOSUM62 Score"
]
col_widths = [9, 10, 10, 10, 10, 15, 15]

for ci, (hdr, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=ci, value=hdr)
    cell.font = HDR_FONT
    cell.fill = HDR_FILL
    cell.alignment = CENTER
    ws.column_dimensions[get_column_letter(ci)].width = w

ws.row_dimensions[1].height = 22
ws.freeze_panes = "A2"

for ri, row in enumerate(mapping, 2):
    h_pos  = row['h_pos']  if row['h_pos']  else ""
    h_aa   = row['h_aa']   if row['h_aa']   else "—"
    m_pos  = row['m_pos']  if row['m_pos']  else ""
    m_aa   = row['m_aa']   if row['m_aa']   else "—"
    blsm   = row['blosum62'] if row['blosum62'] is not None else ""

    if row['identity']:
        cons_label = "Identical"
        row_fill   = IDENT_FILL
    elif row['conservative']:
        cons_label = "Conservative"
        row_fill   = CONS_FILL
    elif not row['h_aa'] or not row['m_aa']:
        cons_label = "Gap"
        row_fill   = GAP_FILL
    else:
        cons_label = "Non-conservative"
        row_fill   = ALT_FILL if (ri % 2 == 0) else NONE_FILL

    vals = [row['aln_col'], h_pos, h_aa, m_pos, m_aa, cons_label, blsm]
    for ci, val in enumerate(vals, 1):
        cell = ws.cell(row=ri, column=ci, value=val)
        cell.font = BODY_FONT
        cell.alignment = CENTER
        cell.fill = row_fill

# ─ Sheet 2: Summary ─
ws2 = wb.create_sheet("Summary")
summary_rows = [
    ("Metric", "Value"),
    ("Human protein", "Interferon lambda-2 (Q8IZJ0, Homo sapiens)"),
    ("Mouse protein",  "Interferon lambda-2 (Q4VK74, Mus musculus)"),
    ("Human length (aa)", len(HUMAN)),
    ("Mouse length (aa)", len(MOUSE)),
    ("Alignment length", aln_len),
    ("Alignment method", "Needleman-Wunsch (global), BLOSUM62"),
    ("Gap open penalty", -10),
    ("Gap extend penalty", -0.5),
    ("Identical residues", ident),
    ("Aligned pairs", len(paired)),
    ("% Identity", f"{pct_id:.1f}%"),
    ("Conservative substitutions", cons),
    ("% Similarity (id + conservative)", f"{pct_sim:.1f}%"),
    ("Gap columns", gaps),
]

ws2.column_dimensions["A"].width = 38
ws2.column_dimensions["B"].width = 38

for ri, (k, v) in enumerate(summary_rows, 1):
    ca = ws2.cell(row=ri, column=1, value=k)
    cb = ws2.cell(row=ri, column=2, value=v)
    if ri == 1:
        for c in (ca, cb):
            c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = CENTER
    else:
        for c in (ca, cb):
            c.font = BODY_FONT
            c.alignment = Alignment(horizontal="left", vertical="center")
        if ri % 2 == 0:
            ca.fill = ALT_FILL; cb.fill = ALT_FILL

wb.save(OUT_XLSX)
print(f"Saved Excel: {OUT_XLSX}")

# ── HTML alignment view ───────────────────────────────────────────────────────
OUT_HTML = "/sessions/affectionate-zealous-hypatia/mnt/outputs/IFNL2_human_mouse_alignment.html"

def classify(h, m):
    if h == '-' or m == '-':
        return 'gap'
    if h.upper() == m.upper():
        return 'identical'
    if blosum_score(h, m) > 0:
        return 'conservative'
    return 'noncons'

BLOCK = 60
aln_len_total = len(aln_human)

blocks_html = []
h_counter, m_counter = 0, 0

for start in range(0, aln_len_total, BLOCK):
    block_h = aln_human[start:start+BLOCK]
    block_m = aln_mouse[start:start+BLOCK]

    # Count positions
    h_start_num = sum(1 for c in aln_human[:start] if c != '-') + 1
    m_start_num = sum(1 for c in aln_mouse[:start] if c != '-') + 1

    h_end_num = sum(1 for c in aln_human[:start+len(block_h)] if c != '-')
    m_end_num = sum(1 for c in aln_mouse[:start+len(block_m)] if c != '-')

    # Build per-column spans — human, consensus, mouse all use same <span class="aa">
    # so widths are guaranteed to match exactly.
    h_cells = ''
    c_cells = ''
    m_cells = ''
    for h, m in zip(block_h, block_m):
        cls = classify(h, m)
        if cls == 'identical':      csym, ccls = '|', 'cons-ident'
        elif cls == 'conservative': csym, ccls = '+', 'cons-cons'
        else:                       csym, ccls = '&nbsp;', 'cons-other'
        h_cells += f'<span class="aa {cls}">{h}</span>'
        c_cells += f'<span class="aa {ccls}">{csym}</span>'
        m_cells += f'<span class="aa {cls}">{m}</span>'

    h_label = f"Human {h_start_num:>4}–{h_end_num:<4}"
    m_label = f"Mouse {m_start_num:>4}–{m_end_num:<4}"

    blocks_html.append(f"""
<div class="block">
  <div class="aln-row"><span class="label">{h_label}</span><span class="seq">{h_cells}</span></div>
  <div class="aln-row"><span class="label"></span><span class="seq">{c_cells}</span></div>
  <div class="aln-row"><span class="label">{m_label}</span><span class="seq">{m_cells}</span></div>
</div>""")

html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>IFN-Lambda 2: Human vs Mouse Alignment</title>
<style>
  body {{ font-family: Arial, sans-serif; background: #f8f9fa; margin: 0; padding: 20px; }}
  h1 {{ color: #2E4057; }}
  .summary {{ background: #fff; border: 1px solid #dee2e6; border-radius: 6px; padding: 16px; margin-bottom: 24px; display: flex; gap: 40px; flex-wrap: wrap; }}
  .stat {{ text-align: center; }}
  .stat .val {{ font-size: 2em; font-weight: bold; color: #2E4057; }}
  .stat .lbl {{ font-size: 0.85em; color: #6c757d; }}
  .legend {{ display: flex; gap: 16px; margin-bottom: 16px; flex-wrap: wrap; }}
  .leg {{ display: flex; align-items: center; gap: 6px; font-size: 0.85em; }}
  .leg-box {{ width: 14px; height: 14px; border-radius: 3px; }}
  .alignment {{ font-family: 'Courier New', monospace; font-size: 13px; line-height: 1.8; }}
  .block {{ background: #fff; border: 1px solid #dee2e6; border-radius: 6px; padding: 10px 16px; margin-bottom: 12px; }}
  .aln-row {{ display: flex; align-items: center; }}
  .label {{ width: 140px; color: #495057; font-size: 12px; flex-shrink: 0; }}
  .seq {{ letter-spacing: 1px; }}
  .cons-seq {{ color: #555; }}
  .aa {{ display: inline-block; width: 1ch; text-align: center; }}
  .aa.identical    {{ background: #d4edda; color: #155724; }}
  .aa.conservative {{ background: #fff3cd; color: #856404; }}
  .aa.noncons      {{ background: #f8d7da; color: #721c24; }}
  .aa.gap          {{ background: #f0f0f0; color: #6c757d; }}
  .aa.cons-ident  {{ background: transparent; color: #155724; font-weight: bold; }}
  .aa.cons-cons   {{ background: transparent; color: #856404; font-weight: bold; }}
  .aa.cons-other  {{ background: transparent; color: transparent; }}
</style>
</head>
<body>
<h1>IFN-Lambda 2: Human (Q8IZJ0) vs Mouse (Q4VK74) — Global Alignment</h1>
<div class="summary">
  <div class="stat"><div class="val">{len(HUMAN)}</div><div class="lbl">Human length (aa)</div></div>
  <div class="stat"><div class="val">{len(MOUSE)}</div><div class="lbl">Mouse length (aa)</div></div>
  <div class="stat"><div class="val">{aln_len}</div><div class="lbl">Alignment length</div></div>
  <div class="stat"><div class="val">{pct_id:.1f}%</div><div class="lbl">Identity</div></div>
  <div class="stat"><div class="val">{pct_sim:.1f}%</div><div class="lbl">Similarity</div></div>
  <div class="stat"><div class="val">{gaps}</div><div class="lbl">Gap columns</div></div>
</div>
<div class="legend">
  <div class="leg"><div class="leg-box" style="background:#d4edda;"></div> Identical</div>
  <div class="leg"><div class="leg-box" style="background:#fff3cd;"></div> Conservative (BLOSUM62 &gt; 0)</div>
  <div class="leg"><div class="leg-box" style="background:#f8d7da;"></div> Non-conservative</div>
  <div class="leg"><div class="leg-box" style="background:#f0f0f0;"></div> Gap</div>
  <div class="leg" style="margin-left:20px;"><b>|</b>&nbsp;= identical&nbsp;&nbsp;<b>+</b>&nbsp;= conservative&nbsp;&nbsp;<b>&nbsp;</b>&nbsp;= non-cons/gap</div>
</div>
<div class="alignment">
{''.join(blocks_html)}
</div>
<p style="color:#6c757d;font-size:12px;margin-top:24px;">
  Method: Needleman-Wunsch global alignment · Scoring: BLOSUM62 · Gap open: –10 · Gap extend: –0.5
</p>
</body>
</html>"""

with open(OUT_HTML, "w") as f:
    f.write(html)
print(f"Saved HTML: {OUT_HTML}")
